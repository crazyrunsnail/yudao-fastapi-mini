import io
from typing import List, Any, Dict, Type, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pydantic import BaseModel

class ExcelUtils:
    @staticmethod
    def write(data: List[Any], fields: Dict[str, str], sheet_name: str = "Sheet1") -> bytes:
        """
        Write data to an Excel file in memory.
        :param data: List of objects or dictionaries to write.
        :param fields: Dictionary mapping field names (keys) to column headers (values).
        :param sheet_name: Name of the sheet.
        :return: Excel file content as bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write header
        headers = list(fields.values())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write data
        field_names = list(fields.keys())
        for row_idx, item in enumerate(data, 2):
            for col_idx, field in enumerate(field_names, 1):
                value = None
                if isinstance(item, dict):
                    value = item.get(field)
                elif isinstance(item, BaseModel):
                    value = getattr(item, field, None)
                else:
                    value = getattr(item, field, None)
                
                # Format value if needed (e.g. datetime)
                if hasattr(value, "isoformat"):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column width
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def read(file_content: bytes, schema: Type[BaseModel], fields: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        Read data from an Excel file.
        :param file_content: Excel file content as bytes.
        :param schema: Pydantic model for validation (metadata only here).
        :param fields: Dictionary mapping column headers (keys) to field names (values).
        :return: List of dictionaries representing the rows.
        """
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        header_to_field = {header: fields.get(header) for header in headers if header in fields}
        
        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for idx, value in enumerate(row):
                header = headers[idx]
                field = header_to_field.get(header)
                if field:
                    row_data[field] = value
            if any(row_data.values()): # Skip empty rows
                result.append(row_data)
        
        return result
